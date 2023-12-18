import os
import sqlite3

import openpyxl
import pandas as pd

from loger import get_logger

logger = get_logger(__name__)
#
# def process_excel_file():
#     input_excel_file = 'bd.xlsx'
#     workbook = openpyxl.load_workbook(input_excel_file)
#     worksheet = workbook.active
#     max_row = worksheet.max_row
#     for row in range(1, max_row + 1):
#         cell = worksheet.cell(row=row, column=4)
#         if cell.value:
#             cell.value = " ".join(word_shortener(str(cell.value)))
#
#     # Сохраните изменения в новом файле
#     output_excel_file = 'bd.xlsx'
#     workbook.save(output_excel_file)



def bd_create():
    """
    Создает базу данных с Excel файла
    :return: ничего не возвращает
    """
    conn = sqlite3.connect('horse_database.db')
    cursor = conn.cursor()
    cursor.execute(
        '''
        CREATE TABLE IF NOT EXISTS my_table (
            id INTEGER PRIMARY KEY,
            name TEXT
        )
        '''
    )
    cursor.close()
    # Загрузка данных из Excel
    df = pd.read_excel('bd.xlsx')
    df = df.apply(lambda x: x.lower() if isinstance(x, str) else x)

    # Загрузка данных в базу данных SQLite
    df.to_sql('my_table', conn, if_exists='replace', index=False)
    conn.commit()
    conn.close()
    logger.debug('БАЗА ДАНЫХ СОЗДАННА')


def delete_bd():
    if os.path.exists('horse_database.db'):
        os.remove('horse_database.db')
        logger.debug("База данных удалена.")
    else:
        logger.debug("База данных не существует.")



def remove_empty_elements(input_list):
    processed_list = [element for element in input_list if element.strip() != '']
    return processed_list



def benchmark(func):
    import time
    def wrapper(*args, **kwargs):
        start = time.time()
        return_value = func(*args, **kwargs)
        end = time.time()
        print(f'[{func.__name__}] Время выполнения: {end - start} секунд.')
        return return_value

    return wrapper

